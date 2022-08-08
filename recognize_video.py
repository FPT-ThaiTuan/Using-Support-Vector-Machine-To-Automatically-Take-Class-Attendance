# import libraries
import os
from copyreg import pickle

import cv2
import imutils
import time
import pickle
import numpy as np
from imutils.video import FPS
from imutils.video import VideoStream
from openpyxl import load_workbook

# Initialize excel file
attendanceFile = load_workbook(filename="take_attendance/AIP391_AI1601_Students.xlsx")
sheet = attendanceFile.active
# attendanceFile = pd.read_excel(r'take_attendance/Attendance.xlsx')
# studentCode = pd.DataFrame(attendanceFile)['RollNumber']
# print(studentCode)
attendedStudent = []

# load serialized face detector
print("Loading Face Detector...")
protoPath = "face_detection_model/deploy.prototxt"
modelPath = "face_detection_model/res10_300x300_ssd_iter_140000.caffemodel"
detector = cv2.dnn.readNetFromCaffe(protoPath, modelPath)

# load serialized face embedding model
print("Loading Face Recognizer...")
embedder = cv2.dnn.readNetFromTorch("openface_nn4.small2.v1.t7")

# load the actual face recognition model along with the label encoder
# a = open("output/recognizer.pickle", "rb")
# recognizer = pickle.loads(a)
recognizer = pickle.loads(open("output/recognizer.pickle", "rb").read())

# a = open("output/le.pickle","rb").readlines()
# a = map(lambda x:x.replace("\r\n","\n"),a)
# with open("output/le.pickle","wb") as j: #write back to file in binary mode
#     for i in a:
#         j.write(i)
# pickle.load(open("output/le.pickle","rb"))
# b = open("output/le.pickle", "rb")
# le = pickle.loads(b)
le = pickle.loads(open("output/le.pickle", "rb").read())

# initialize the video stream, then allow the camera sensor to warm up
print("Starting Video Stream...")
vs = VideoStream(src=0).start()
time.sleep(2.0)

# start the FPS throughput estimator
fps = FPS().start()

# loop over frames from the video file stream
while True:
	# grab the frame from the threaded video stream
	frame = vs.read()

	# resize the frame to have a width of 600 pixels (while maintaining the aspect ratio), and then grab the image dimensions
	frame = imutils.resize(frame, width=600)
	(h, w) = frame.shape[:2]

	# construct a blob from the image
	imageBlob = cv2.dnn.blobFromImage(
		cv2.resize(frame, (300, 300)), 1.0, (300, 300),
		(104.0, 177.0, 123.0), swapRB=False, crop=False)

	# apply OpenCV's deep learning-based face detector to localize faces in the input image
	detector.setInput(imageBlob)
	detections = detector.forward()

	# loop over the detections
	for i in range(0, detections.shape[2]):
		# extract the confidence (i.e., probability) associated with the prediction
		confidence = detections[0, 0, i, 2]

		# filter out weak detections
		if confidence > 0.5:
			# compute the (x, y)-coordinates of the bounding box for the face
			box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
			(startX, startY, endX, endY) = box.astype("int")

			# extract the face ROI
			face = frame[startY:endY, startX:endX]
			(fH, fW) = face.shape[:2]

			# ensure the face width and height are sufficiently large
			if fW < 20 or fH < 20:
				continue

			# construct a blob for the face ROI, then pass the blob through our face embedding model to obtain the 128-d quantification of the face
			faceBlob = cv2.dnn.blobFromImage(face, 1.0 / 255,
				(96, 96), (0, 0, 0), swapRB=True, crop=False)
			embedder.setInput(faceBlob)
			vec = embedder.forward()

			# perform classification to recognize the face
			preds = recognizer.predict_proba(vec)[0]
			j = np.argmax(preds)
			proba = preds[j]
			name = le.classes_[j]

			# draw the bounding box of the face along with the associated probability
			text = "{}: {:.2f}%".format(name, proba * 100)
			y = startY - 10 if startY - 10 > 10 else startY + 10
			cv2.rectangle(frame, (startX, startY), (endX, endY),
				(0, 0, 255), 2)
			cv2.putText(frame, text, (startX, y),
				cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 255), 2)
			if(proba > 0.85): attendedStudent.append(name)
	# update the FPS counter
	fps.update()

	# show the output frame
	cv2.imshow("Frame", frame)
	key = cv2.waitKey(1) & 0xFF

	# if the `q` key was pressed, break from the loop
	if key == ord("q"):
		break

# stop the timer and display FPS information
fps.stop()
print("Elasped time: {:.2f}".format(fps.elapsed()))
print("Approx. FPS: {:.2f}".format(fps.fps()))

# cleanup
cv2.destroyAllWindows()
vs.stop()

attendedStudent = list(set(attendedStudent))
attendedStudent = np.char.upper(attendedStudent)

# Get attendance
for i in range(sheet.max_row):
	if sheet.cell(i + 2, 1).value in attendedStudent:
		sheet.cell(i + 2, 8).value = "Attend"
attendanceFile.save("take_attendance/AIP391_AI1601_Students.xlsx")